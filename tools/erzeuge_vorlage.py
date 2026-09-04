"""Erzeugt eine veröffentlichbare Dashboard-Vorlage aus einer echten Mappe.

Die Vorlage behält das Raster, Formeln, Merges und die Tabellenformatierung.
Sie enthält keine Bestands-, Bestell- oder Empfangsdaten. Dieses Werkzeug ist
absichtlich kein Teil des normalen Starts; es wird nur beim Aktualisieren der
Vorlage benutzt.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from bestand.core import parse_grid, resolve_anchor
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import range_boundaries

WURZEL = Path(__file__).resolve().parent.parent
STANDARD_EINGABE = (
    WURZEL.parent / "sba-bestand" / "bestand" / "Bestand- und Nachbestellungsliste 2026.xlsx"
)
STANDARD_AUSGABE = WURZEL / "vorlage" / "Bestand- und Nachbestellungsliste 2026.xlsx"


def _leere_rasterdaten(ws) -> None:
    """Löscht nur die Werte, die das Raster als veränderliche Zahlen kennt."""
    grid = parse_grid(ws)
    refs = set()
    for stand_row in grid.stand_rows:
        anchor_row, anchor_col = resolve_anchor(ws, stand_row, 2)
        refs.add(ws.cell(anchor_row, anchor_col).coordinate)
    for eintrag in grid.entries:
        refs.update(eintrag.angemeldet_refs)
        refs.add(eintrag.slots["bestand"].ref)
        refs.add(eintrag.slots["bestellt"].ref)
    for ref in refs:
        ws[ref].value = None


def _leere_tabellenkoerper(ws) -> None:
    """Entfernt alle eingegebenen Tabellenwerte, aber behält berechnete Spalten."""
    for tabelle in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(tabelle.ref)
        for zeile in ws.iter_rows(
            min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for zelle in zeile:
                if zelle.data_type != "f":
                    zelle.value = None


def bereinige(eingabe: Path, ausgabe: Path) -> None:
    wb = load_workbook(eingabe, data_only=False)
    raster = wb["Bestand- und Nachbestellung"]
    _leere_rasterdaten(raster)

    for ws in wb.worksheets:
        _leere_tabellenkoerper(ws)
        for zeile in ws.iter_rows():
            for zelle in zeile:
                if isinstance(zelle, MergedCell):
                    continue
                zelle.comment = None
                zelle.hyperlink = None

    eigenschaften = wb.properties
    eigenschaften.creator = "sba-dashboard"
    eigenschaften.lastModifiedBy = "sba-dashboard"
    eigenschaften.title = "Bestand- und Nachbestellungsliste 2026 (Vorlage)"
    eigenschaften.subject = "Leere Vorlage für das Schulbuchausleihe-Dashboard"
    eigenschaften.description = "Enthält die Excel-Struktur ohne Arbeitsdaten."
    eigenschaften.keywords = ""

    ausgabe.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ausgabe)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Bereinigt eine echte Mappe zur Git-Vorlage.")
    parser.add_argument("--eingabe", type=Path, default=STANDARD_EINGABE)
    parser.add_argument("--ausgabe", type=Path, default=STANDARD_AUSGABE)
    argumente = parser.parse_args(argv)

    if not argumente.eingabe.is_file():
        parser.error(f"Eingabedatei fehlt: {argumente.eingabe}")
    bereinige(argumente.eingabe, argumente.ausgabe)
    print(f"Vorlage geschrieben: {argumente.ausgabe}")
    return 0


if __name__ == "__main__":  # pragma: no cover - Konsoleneinstieg
    raise SystemExit(main())
