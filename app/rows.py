"""Vom Excel-Raster zur flachen Liste - eine Zeile je Buch und Jahrgangsband.

Das Raster ist 62 Spalten breit und für Menschen gebaut: Fachblöcke nebeneinander,
Jahrgänge untereinander, Mehrjahresbände als Zellverbund. Die Oberfläche zeigt
stattdessen eine gewöhnliche Liste - filterbar, sortierbar, ohne Scrollen nach rechts.

Zwei Regeln bestimmen die Umrechnung:

* **Angemeldet wird summiert.** In einem Mehrjahresband (K3:K4 = Jg 5-6) steht der
  Bestand einmal, die Anmeldungen aber je Jahrgang einzeln (J3=94, J4=73). Die
  Excel-Formel macht daraus ``=J3+J4-K3-L3``; hier wird genauso summiert.
* **"zu bestellen" wird gerechnet, nie gelesen.** Die Spalte enthält Formeltext.
  Ohne laufendes Excel hat sie keinen Wert, deshalb rechnet Python selbst.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from bestand.core import Grid, GridEntry, parse_grid
from openpyxl.worksheet.worksheet import Worksheet

from . import cache as cache_modul
from .cache import Cache
from .excel import Dateizustand, lade_mappe, raster_blatt
from .settings import Einstellungen


@dataclass(frozen=True)
class Zeile:
    """Eine Zeile der Tabelle. ``key`` ist der stabile Schlüssel des Rasters."""
    key: str
    fach: str
    jahrgang: str
    jahrgaenge: tuple[int, ...]
    titel: str | None
    isbn: str | None
    angemeldet: int | None
    bestand: int | None
    bestellt: int | None
    zu_bestellen: int | None
    bestand_ref: str
    bestellt_ref: str | None
    angemeldet_refs: tuple[str, ...]

    @property
    def bedarf(self) -> bool:
        return (self.zu_bestellen or 0) > 0

    def als_dict(self) -> dict[str, Any]:
        return {
            "key": self.key,
            "fach": self.fach,
            "jahrgang": self.jahrgang,
            "titel": self.titel,
            "isbn": self.isbn,
            "angemeldet": self.angemeldet,
            "bestand": self.bestand,
            "bestellt": self.bestellt,
            "zu_bestellen": self.zu_bestellen,
            "bestand_ref": self.bestand_ref,
            "bestellt_ref": self.bestellt_ref,
            # Mehrzahl, weil ein Mehrjahresband den Bestand einmal führt, die
            # Anmeldungen aber je Jahrgang einzeln. Steht seit 2026-09-05 mit in
            # der Antwort, damit die Oberfläche nach einem Abruf auch diese
            # Spalte hervorheben kann - die geänderten Zellen kommen vom Server
            # als Zellbezüge, und ohne die Anmeldungs-Bezüge bliebe die
            # meistgeänderte Spalte der Tabelle als einzige stumm.
            "angemeldet_refs": list(self.angemeldet_refs),
        }


def _zahl(wert: Any) -> int | None:
    """Zellwert als ganze Zahl - Formeltext und Fremdformate zählen als leer."""
    if wert is None or isinstance(wert, str):
        return None
    if isinstance(wert, bool):
        return None
    try:
        return int(wert)
    except (TypeError, ValueError):
        return None


def _summe_angemeldet(ws: Worksheet, refs: tuple[str, ...]) -> int | None:
    """Summe über die Jahrgangszellen; None, wenn keine einzige gefüllt ist."""
    werte = [_zahl(ws[ref].value) for ref in refs]
    gefuellt = [w for w in werte if w is not None]
    return sum(gefuellt) if gefuellt else None


def zeile_aus_eintrag(ws: Worksheet, eintrag: GridEntry, cache: Cache) -> Zeile:
    bestand_ref = eintrag.slots["bestand"].ref
    bestellt_slot = eintrag.slots.get("bestellt")
    bestellt_ref = bestellt_slot.ref if bestellt_slot else None

    angemeldet = _summe_angemeldet(ws, eintrag.angemeldet_refs)
    bestand = _zahl(ws[bestand_ref].value)
    bestellt = _zahl(ws[bestellt_ref].value) if bestellt_ref else None
    # Bedarf immer selbst rechnen; die Formelspalte bleibt unberührt.
    zu_bestellen = None if angemeldet is None else angemeldet - (bestand or 0) - (bestellt or 0)

    eintrag_cache = cache.get(eintrag.key)
    return Zeile(
        key=eintrag.key,
        fach=eintrag.fach_label,
        jahrgang=eintrag.grade_label,
        jahrgaenge=eintrag.grades,
        titel=eintrag_cache.titel,
        isbn=eintrag_cache.isbn,
        angemeldet=angemeldet,
        bestand=bestand,
        bestellt=bestellt,
        zu_bestellen=zu_bestellen,
        bestand_ref=bestand_ref,
        bestellt_ref=bestellt_ref,
        angemeldet_refs=eintrag.angemeldet_refs,
    )


def baue_zeilen(ws: Worksheet, grid: Grid, cache: Cache | None = None) -> list[Zeile]:
    """Alle Zeilen des Rasters. Sperrflächen sind im Grid bereits ausgesondert."""
    cache = cache or Cache()
    return [zeile_aus_eintrag(ws, eintrag, cache) for eintrag in grid.entries]


@dataclass(frozen=True)
class Tabellenstand:
    """Alles, was ein Seitenaufbau über die Mappe wissen muss - in einem Stück.

    Bis 2026-09-05 gab ``lies_tabelle`` ein untypisiertes Vierertupel zurück,
    dessen Leerfall ``(None, [], None, None)`` hieß. Jeder Aufrufer zerlegte es
    erst in vier Namen und prüfte danach einen davon auf ``None`` - drei Zeilen
    Zeremonie je Route, und die drei ``None``-Felder waren für einen Typprüfer
    (und für den Leser) nicht von echten Werten zu unterscheiden.

    Jetzt gilt: entweder es gibt eine Mappe, dann sind **alle** Felder gesetzt,
    oder ``lies_tabelle`` gibt ``None`` zurück. Der Leerfall steht damit einmal
    im Rückgabetyp statt viermal im Inhalt.
    """

    pfad: Path
    zeilen: list[Zeile]
    zustand: Dateizustand
    cache: Cache

    @property
    def bedarf_gesamt(self) -> int:
        """Summe der offenen Stückzahlen - dieselbe Zahl, die der Browser nachrechnet."""
        return sum(z.zu_bestellen or 0 for z in self.zeilen if z.bedarf)


def lies_tabelle(einstellungen: Einstellungen) -> Tabellenstand | None:
    """Mappe frisch laden, Raster parsen, Anzeigezeilen bauen.

    ``None`` heißt: keiner der eingetragenen Pfade existiert. Das ist der
    Normalzustand vor der Ersteinrichtung und kein Fehler - die Startseite
    zeigt dann die Einrichtungsseite, die API antwortet mit 503 und der Liste
    der geprüften Pfade.

    Geladen wird ohne Schreibsperre. Das ist Absicht: ein Leser soll nie auf
    einen Schreiber warten müssen (siehe ``docs/architektur.md``).
    """
    pfad = einstellungen.excel_pfad()
    if pfad is None:
        return None
    wb = lade_mappe(pfad)
    ws = raster_blatt(wb, einstellungen.blatt_raster)
    grid = parse_grid(ws)
    cache = cache_modul.laden(pfad)
    return Tabellenstand(pfad=pfad, zeilen=baue_zeilen(ws, grid, cache),
                         zustand=Dateizustand.von(pfad), cache=cache)
