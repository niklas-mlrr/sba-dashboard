"""Zugriff auf die Arbeitsmappe - immer frisch laden, nie zwischen Anfragen halten.

Die Mappe liegt auf einem Netzlaufwerk und kann jederzeit in Excel geöffnet sein.
Ein im Speicher gehaltenes Workbook würde still veralten, deshalb lädt jede
Anfrage neu und merkt sich nur die Änderungszeit.

``data_only=False`` ist Pflicht: die Spalte "zu bestellen" enthält echte Formeln.
Mit ``data_only=True`` gelesen und wieder gespeichert wären sie unwiderruflich
durch ihren letzten berechneten Wert ersetzt.

Der Schreibpfad hat drei Schutzschichten, die alle drei nötig sind:

* **Kein freier Zellbezug.** ``/api/cell`` bekommt den Schlüssel einer Zeile,
  nie eine Referenz wie ``"K3"``. Der Schlüssel wird gegen das *frisch* geparste
  Raster aufgelöst; steht die Zeile inzwischen woanders, schreibt der Server
  trotzdem in die richtige Zelle - oder gar nicht.
* **Optimistisches Sperren.** Der Browser schickt die Änderungszeit mit, die er
  gesehen hat. Weicht sie ab, hat jemand anderes (oder der Abruf) die Mappe
  angefasst und die Änderung wird abgelehnt statt sie zu überschreiben.
* **Atomar speichern.** Erst vollständig in eine Nachbardatei schreiben, dann
  ``os.replace``. Ein Abbruch mittendrin lässt die alte Mappe unberührt.
  ``atomic_save_workbook`` liegt dafür in ``bestand.core`` — dort, wo auch die
  CLI der Bestandsliste sie benutzt.

Dazu kommt :func:`validiere_excel_mappe`: die Prüfung einer **neu ausgewählten**
Datei, bevor ihr Pfad in die Benutzerkonfiguration wandert. Sie steht seit dem
2026-09-05 hier und nicht mehr in ``app/main.py`` - sie enthält keine Zeile
FastAPI, sondern beantwortet allein die Frage, ob eine Datei als Bestandsmappe
taugt.

Keine Funktion in dieser Datei weiß etwas über HTTP. Die Ausnahmen sagen, *was*
schiefging; welcher Statuscode daraus wird, entscheidet ``app/fehler.py`` an
einer Stelle für die ganze Anwendung.
"""
from __future__ import annotations

import os
import sys
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from math import isfinite
from pathlib import Path
from typing import IO, Iterator
from zipfile import BadZipFile

# Aus bestand.core, nicht mehr aus dem IServ-Client: dauerhaftes Speichern
# einer Mappe kennt weder IServ noch HTTP (siehe docs/verteilung.md).
from bestand.core import GridEntry, atomic_save_workbook, parse_grid
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

# Nur diese eine Spalte ist von Hand änderbar.
#
# "Angemeldet" und "Bestand" kommen aus IServ und werden bei jedem Abruf
# überschrieben - eine Eingabe dort hielte höchstens bis zum nächsten Abruf und
# sähe bis dahin aus wie eine Zahl, auf die man sich verlassen kann.
# "zu bestellen" ist eine Formel, die ein Schreibzugriff plattmachen würde.
# Bleibt "Bestellt": es kommt nicht aus IServ, sondern aus dem Blatt 'bestellt'
# derselben Mappe, und was dort (noch) nicht steht, lässt der Abruf seit
# 2026-09-05 unangetastet stehen (``bestand.core.update.apply_snapshot``).
# Genau deshalb ist es die einzige Spalte, in der eine Eingabe von Hand Bestand
# hat - und die einzige, die das Dashboard schreiben lässt.
SCHREIBBARE_SPALTEN = ("bestellt",)


def erlaubte_spalten_satz() -> str:
    """Der Satz, der die schreibbaren Spalten aufzählt - Einzahl wie Mehrzahl.

    Steht hier und nicht zweimal ausgeschrieben in ``app/excel.py`` und
    ``app/modelle.py``: seit ``SCHREIBBARE_SPALTEN`` nur noch einen Eintrag hat,
    ergab das dortige ``' und '.join(...)`` den Satz "Erlaubt sind bestand und
    bestellt." minus die Hälfte - also "Erlaubt sind bestellt.". Die Zahl der
    Spalten darf sich ändern, ohne dass jemand daran denken muss.
    """
    namen = [f"'{spalte}'" for spalte in SCHREIBBARE_SPALTEN]
    if len(namen) == 1:
        return f"Änderbar ist nur die Spalte {namen[0]}."
    return "Änderbar sind nur die Spalten " + " und ".join(
        [", ".join(namen[:-1]), namen[-1]]
    ) + "."


class ExcelFehlt(FileNotFoundError):
    """Die Arbeitsmappe liegt unter keinem der eingetragenen Pfade."""


class BlattFehlt(KeyError):
    """Das erwartete Tabellenblatt fehlt in der Mappe."""


class UngueltigeAenderung(ValueError):
    """Schlüssel, Spalte oder Wert sind nicht schreibbar - führt zu HTTP 400."""


class Konflikt(RuntimeError):
    """Die Mappe wurde seit dem Laden geändert - führt zu HTTP 409."""

    def __init__(self, meldung: str, aktuelle_mtime: float) -> None:
        super().__init__(meldung)
        self.aktuelle_mtime = aktuelle_mtime


class Gesperrt(RuntimeError):
    """Die Mappe ist in Excel geöffnet - führt zu HTTP 423."""

    def __init__(self, meldung: str, benutzer: str | None = None) -> None:
        super().__init__(meldung)
        self.benutzer = benutzer


class MappeUngeeignet(ValueError):
    """Eine *neu ausgewählte* Datei taugt nicht als Bestandsmappe - HTTP 400.

    Abgegrenzt von :class:`BlattFehlt`: dort fehlt das Blatt in der Mappe, mit
    der das Dashboard bereits arbeitet - ein Serverzustand (HTTP 500), an dem
    die Lehrkraft im Moment der Anfrage nichts falsch gemacht hat. Hier
    dagegen liegt der Fehler in der gerade **eingegebenen** Auswahl, und die
    Antwort ist eine Bitte, eine andere Datei zu wählen. Erst diese Trennung
    macht eine einzige, zentrale Fehlerabbildung möglich (``app/fehler.py``);
    vorher fing die Einrichtungsroute dieselben Ausnahmeklassen mit einem
    anderen Status ab als die Leseroute.
    """


_prozesssperren_guard = threading.Lock()
_prozesssperren: dict[Path, threading.Lock] = {}


def _prozesssperre(pfad: Path) -> threading.Lock:
    """Das lokale Schloss für eine Mappe, unabhängig vom aktuellen Arbeitsordner."""
    kanonisch = pfad.resolve()
    with _prozesssperren_guard:
        return _prozesssperren.setdefault(kanonisch, threading.Lock())


def _sperrpfad(pfad: Path) -> Path:
    """Neben der Mappe liegende, dauerhafte Datei für SMB-taugliche Dateisperren."""
    return pfad.with_name(f".{pfad.name}.sba-dashboard.lock")


def _versuche_dateisperre(datei: IO[bytes]) -> bool:
    """Versucht eine exklusive Byte-Sperre, ohne auf eine Plattform festgelegt zu sein.

    ``sys.platform`` statt ``os.name``: dieselbe Frage, aber die Schreibweise,
    die auch ``app/paths.py`` benutzt (Begründung dort) - und die einzige, die
    ein Typprüfer versteht. Mit ``os.name == "nt"`` meldete mypy auf Linux vier
    Fehler in den ``msvcrt``-Zeilen, weil dieses Modul dort nicht existiert;
    mit ``sys.platform`` kennt es den Zweig als plattformgebunden und prüft je
    Plattform nur den Zweig, den es dort auch gibt.
    """
    if sys.platform == "win32":  # pragma: no cover - auf dem Windows-Schulrechner ausgeführt
        import msvcrt

        datei.seek(0)
        try:
            msvcrt.locking(datei.fileno(), msvcrt.LK_NBLCK, 1)
        except OSError:
            return False
        return True

    import fcntl

    try:
        fcntl.flock(datei.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        return False
    return True


def _loese_dateisperre(datei: IO[bytes]) -> None:
    if sys.platform == "win32":  # pragma: no cover - auf dem Windows-Schulrechner ausgeführt
        import msvcrt

        datei.seek(0)
        msvcrt.locking(datei.fileno(), msvcrt.LK_UNLCK, 1)
    else:
        import fcntl

        fcntl.flock(datei.fileno(), fcntl.LOCK_UN)


@contextmanager
def arbeitsmappe_sperren(pfad: Path, *, wartezeit: float = 30.0) -> Iterator[None]:
    """Serialisiert den vollständigen Lese-Ändere-Speichere-Zugriff auf ``pfad``.

    Das Python-Schloss schützt Threads des lokalen Dashboard-Prozesses. Die
    dauerhafte Nachbardatei hält zusätzlich eine exklusive Betriebssystem-
    Dateisperre und koordiniert damit weitere Dashboard-Prozesse bzw. Rechner,
    sofern das SMB-Laufwerk Dateisperren unterstützt. Der Inhalt der
    Sperrdatei ist bedeutungslos; sie bleibt bewusst nach dem Entsperren liegen.

    Der Kontext muss *vor* ``lade_mappe`` beginnen und erst *nach*
    ``speichere_mappe`` enden. Nur dann kann kein zweiter Schreiber eine alte
    Mappe laden und später eine neuere Fassung überschreiben.
    """
    if wartezeit < 0:
        raise ValueError("wartezeit darf nicht negativ sein.")
    prozesssperre = _prozesssperre(pfad)
    if not prozesssperre.acquire(timeout=wartezeit):
        raise Gesperrt(
            "Ein anderer Dashboard-Schreibvorgang bearbeitet die Datei noch. "
            "Bitte erneut versuchen."
        )

    datei = None
    gesperrt = False
    try:
        sperrpfad = _sperrpfad(pfad)
        # a+b erzeugt die Datei bei der ersten Benutzung. Ein Byte ist für
        # msvcrt.locking nötig; auf POSIX stört es nicht.
        datei = sperrpfad.open("a+b")
        datei.seek(0, os.SEEK_END)
        if datei.tell() == 0:
            datei.write(b"\0")
            datei.flush()

        ende = time.monotonic() + wartezeit
        while not _versuche_dateisperre(datei):
            if time.monotonic() >= ende:
                raise Gesperrt(
                    "Ein anderes SBA Dashboard bearbeitet die Datei noch. "
                    "Bitte erneut versuchen."
                )
            time.sleep(min(0.05, max(0.0, ende - time.monotonic())))
        gesperrt = True
        yield
    finally:
        if datei is not None:
            if gesperrt:
                _loese_dateisperre(datei)
            datei.close()
        prozesssperre.release()


@dataclass(frozen=True)
class Dateizustand:
    pfad: Path
    mtime: float
    geaendert: datetime

    @classmethod
    def von(cls, pfad: Path) -> "Dateizustand":
        stat = pfad.stat()
        return cls(pfad=pfad, mtime=stat.st_mtime,
                   geaendert=datetime.fromtimestamp(stat.st_mtime))


@dataclass(frozen=True)
class Schreibergebnis:
    ref: str
    # Der Zellwert vor der Änderung, so wie openpyxl ihn hergibt: Zahl, Text,
    # Datum oder None. ``object`` ist hier kein Notausgang, sondern die
    # ehrliche Signatur einer Tabellenzelle - anders als bei ``ws``/``eintrag``,
    # die bis 2026-09-05 nur deshalb ``object`` waren, weil sie mit einem
    # ``None``-Vorgabewert nachträglich angebaut wurden.
    alt: object
    neu: int | None
    zustand: Dateizustand
    backup: Path | None
    # Blatt und Rastereintrag der geschriebenen Zeile. Der Aufrufer baut daraus
    # die aktualisierte Tabellenzeile, ohne die Mappe ein zweites Mal zu laden.
    # Ohne Vorgabewert: ``schreibe_zelle`` ist der einzige Erzeuger und setzt
    # beide immer - ein ``Schreibergebnis`` ohne Blatt hat es nie gegeben.
    ws: Worksheet
    eintrag: GridEntry


def sperrdatei(pfad: Path) -> Path | None:
    """Die ``~$…``-Datei, die Excel beim Öffnen daneben legt - oder None."""
    kandidat = pfad.parent / f"~${pfad.name}"
    return kandidat if kandidat.exists() else None


def _benutzer_aus_sperrdatei(roh: bytes) -> str | None:
    """Der Name aus einer ``~$…``-Datei: Längenbyte, dann der Benutzername.

    Excel unter Windows legt den Namen als UTF-16LE ab Byte 2 ab, ältere und
    Mac-Versionen als 8-Bit-Text ab Byte 1. Beides wird versucht; passt nichts,
    bleibt es beim anonymen "ist geöffnet" - ein geratener Name wäre schlimmer
    als keiner.
    """
    if len(roh) < 2:
        return None
    laenge = roh[0]
    if not 1 <= laenge <= 54:
        return None
    utf16 = roh[2:2 + laenge * 2].decode("utf-16-le", "ignore")
    acht_bit = roh[1:1 + laenge].decode("cp1252", "ignore")
    # Byte 1 entscheidet: in der UTF-16-Fassung ist es das Nullbyte des ersten
    # Zeichens, in der 8-Bit-Fassung schon der erste Buchstabe. Ohne diese Probe
    # liest sich "j.klein" als UTF-16 zu druckbarem CJK-Unsinn - und der wuerde
    # als Name durchgehen.
    kandidaten = (utf16, acht_bit) if roh[1] == 0 else (acht_bit, utf16)
    for text in kandidaten:
        name = text.split("\x00")[0].strip()
        if name and name.isprintable():
            return name
    return None


def sperr_benutzer(pfad: Path) -> str | None:
    """Wer die Mappe geöffnet hat, laut ``~$…``-Datei - oder None."""
    datei = sperrdatei(pfad)
    if datei is None:
        return None
    try:
        return _benutzer_aus_sperrdatei(datei.read_bytes())
    except OSError:
        return None


def sperrmeldung(pfad: Path) -> str:
    """Klartext für HTTP 423, mit Namen wenn die ``~$…``-Datei einen hergibt."""
    benutzer = sperr_benutzer(pfad)
    wer = f" von {benutzer}" if benutzer else ""
    return (f"Die Datei ist gerade in Excel geöffnet{wer}. "
            "Bitte dort schließen und erneut versuchen.")


def lade_mappe(pfad: Path) -> Workbook:
    """Lädt die Mappe mit erhaltenen Formeln."""
    if not pfad.is_file():
        raise ExcelFehlt(f"Excel-Datei nicht gefunden: {pfad}")
    return load_workbook(str(pfad), data_only=False)


def raster_blatt(wb: Workbook, blattname: str) -> Worksheet:
    if blattname not in wb.sheetnames:
        raise BlattFehlt(f"Tabellenblatt {blattname!r} fehlt in der Mappe.")
    return wb[blattname]


# Ohne diese beiden Blätter kann der Abruf nicht laufen: ``bestellt`` liefert die
# bereits bestellten Stückzahlen, ``zu Bestellen`` wird bei jedem Abruf neu
# aufgebaut (siehe app/refresh.py).
ERFORDERLICHE_ZUSATZBLAETTER = ("bestellt", "zu Bestellen")


def validiere_excel_mappe(pfad: Path, blatt_raster: str) -> None:
    """Prüft eine **neu ausgewählte** Datei auf die Struktur, die das Dashboard braucht.

    Läuft vor dem Schreiben der Konfiguration. So kann keine beliebige
    ``.xlsx``-Datei den funktionierenden Pfad in der Benutzerkonfiguration
    verdrängen - und die Lehrkraft erfährt den Grund sofort und nicht erst beim
    nächsten Seitenaufbau.

    Jeder Fehlschlag wird zu :class:`MappeUngeeignet` mit Klartext. Auch das
    fehlende Raster-Blatt: an dieser Stelle ist es keine kaputte Mappe des
    Servers, sondern eine falsch ausgewählte Datei.
    """
    try:
        wb = lade_mappe(pfad)
        ws = raster_blatt(wb, blatt_raster)
        fehlend = [b for b in ERFORDERLICHE_ZUSATZBLAETTER if b not in wb.sheetnames]
        if fehlend:
            raise MappeUngeeignet(
                f"Erforderliche Tabellenblätter fehlen: {', '.join(fehlend)}."
            )
        if not parse_grid(ws).entries:
            raise MappeUngeeignet("Das Tabellenblatt enthält kein lesbares Bestandsraster.")
    except MappeUngeeignet:
        raise
    except BlattFehlt as exc:
        # BlattFehlt erbt von KeyError; str(exc) wäre damit in Anführungszeichen
        # gesetzt (KeyError.__str__ == repr des Arguments). args[0] ist der
        # Klartext, den raster_blatt geschrieben hat.
        raise MappeUngeeignet(str(exc.args[0])) from exc
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError, RuntimeError) as exc:
        raise MappeUngeeignet(f"Die Datei ist keine lesbare Excel-Arbeitsmappe: {exc}") from exc


# ── Schreibpfad ───────────────────────────────────────────────────────────────

def pruefe_wert(roh: object) -> int | None:
    """Erlaubt ist eine ganze Zahl >= 0 oder leer. Sonst nichts.

    Leer heißt bewusst ``None`` und nicht ``0``: eine leere "Bestellt"-Zelle
    bedeutet "nichts bestellt", eine ``0`` bedeutet "nachgesehen, nichts offen".
    Die Mappe unterscheidet das, also unterscheidet es der Server auch.
    """
    if roh is None:
        return None
    if isinstance(roh, str):
        text = roh.strip()
        if not text:
            return None
        try:
            zahl = int(text)
        except ValueError:
            raise UngueltigeAenderung(
                f"{text!r} ist keine Zahl. Erlaubt sind ganze Zahlen ab 0 oder ein leeres Feld."
            ) from None
    elif isinstance(roh, bool):
        # Muss VOR isinstance(roh, int) stehen: bool ist in Python eine
        # Unterklasse von int, ``True`` ginge sonst als Stückzahl 1 durch.
        raise UngueltigeAenderung("Erlaubt sind nur ganze Zahlen ab 0 oder ein leeres Feld.")
    elif isinstance(roh, int):
        zahl = roh
    elif isinstance(roh, float) and roh.is_integer():
        # 5.0 aus JSON ist dieselbe Zahl wie 5; 3.5 und NaN sind es nicht.
        zahl = int(roh)
    else:
        raise UngueltigeAenderung("Erlaubt sind nur ganze Zahlen ab 0 oder ein leeres Feld.")
    if zahl < 0:
        raise UngueltigeAenderung("Eine Stückzahl kann nicht negativ sein.")
    return zahl


def kuerze_backups(backup_dir: Path, behalten: int) -> list[Path]:
    """Löscht die ältesten Backups; gibt die gelöschten Pfade zurück.

    Ein Backup je Änderung summiert sich auf einem Netzlaufwerk schnell. Sortiert
    wird nach Änderungszeit, nicht nach Namen - ein umbenanntes Backup soll
    trotzdem an der richtigen Stelle einsortiert werden.
    """
    if behalten < 0 or not backup_dir.is_dir():
        return []
    dateien = [p for p in backup_dir.iterdir() if p.is_file() and p.suffix == ".xlsx"]
    if len(dateien) <= behalten:
        return []
    dateien.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    geloescht: list[Path] = []
    for pfad in dateien[behalten:]:
        try:
            pfad.unlink()
        except OSError:
            continue
        geloescht.append(pfad)
    return geloescht


def speichere_mappe(wb: Workbook, pfad: Path, *, backups_behalten: int = 30) -> Path | None:
    """Atomar speichern, Backup anlegen, alte Backups kürzen.

    ``PermissionError`` heißt auf Windows praktisch immer: die Datei ist in Excel
    offen. Sie wird in :class:`Gesperrt` übersetzt, damit der Aufrufer nicht
    raten muss, was ein Betriebssystemfehler für die Lehrkraft bedeutet.
    """
    backup_dir = pfad.parent / "backups"
    try:
        backup = atomic_save_workbook(wb, pfad, backup_dir=backup_dir)
    except PermissionError as exc:
        raise Gesperrt(sperrmeldung(pfad), sperr_benutzer(pfad)) from exc
    except OSError as exc:
        # Netzlaufwerk weg, Platte voll, Datei schreibgeschützt - alles Fälle, in
        # denen die alte Mappe dank atomarem Ersetzen unversehrt ist.
        raise Gesperrt(
            f"Die Datei ließ sich nicht speichern: {exc.strerror or exc}. "
            "Die bisherige Fassung ist unverändert.",
            sperr_benutzer(pfad),
        ) from exc
    kuerze_backups(backup_dir, backups_behalten)
    return backup


def schreibe_zelle(
    pfad: Path,
    blattname: str,
    *,
    key: str,
    spalte: str,
    wert: object,
    mtime: float,
    backups_behalten: int = 30,
) -> Schreibergebnis:
    """Setzt eine einzelne Zahl in der Mappe - Schlüssel statt Zellbezug.

    ``mtime`` ist absichtlich Pflicht: sie ist der beim Laden gesehene
    Versionsstand. Der Ablauf ist unter :func:`arbeitsmappe_sperren` bewusst
    "laden, prüfen, schreiben, speichern" in einem Zug. Damit kann weder ein
    zweiter Thread noch ein anderes Dashboard eine alte Mappe laden und später
    eine neuere Fassung überschreiben.
    """
    if spalte not in SCHREIBBARE_SPALTEN:
        raise UngueltigeAenderung(
            f"Die Spalte {spalte!r} ist nicht änderbar. {erlaubte_spalten_satz()}"
        )
    neuer_wert = pruefe_wert(wert)
    if isinstance(mtime, bool) or not isinstance(mtime, (int, float)) or not isfinite(mtime):
        raise UngueltigeAenderung("Es fehlt eine gültige Änderungszeit der geladenen Datei.")

    with arbeitsmappe_sperren(pfad):
        if not pfad.is_file():
            raise ExcelFehlt(f"Excel-Datei nicht gefunden: {pfad}")
        zustand = Dateizustand.von(pfad)
        if abs(zustand.mtime - float(mtime)) > 1e-6:
            raise Konflikt(
                "Die Datei wurde inzwischen geändert. Bitte die Seite neu laden und "
                "die Änderung erneut eintragen.",
                zustand.mtime,
            )

        wb = lade_mappe(pfad)
        ws = raster_blatt(wb, blattname)
        grid = parse_grid(ws)
        eintrag = grid.entry(key)
        if eintrag is None:
            raise UngueltigeAenderung(
                "Diese Zeile gibt es in der Mappe nicht (mehr). Bitte die Seite neu laden."
            )
        slot = eintrag.slots.get(spalte)
        if slot is None:
            raise UngueltigeAenderung(
                f"Für {eintrag.fach_label} Jg. {eintrag.grade_label} gibt es keine Spalte {spalte!r}."
            )

        alt = ws[slot.ref].value
        ws[slot.ref].value = neuer_wert
        backup = speichere_mappe(wb, pfad, backups_behalten=backups_behalten)
        return Schreibergebnis(
            ref=slot.ref, alt=alt, neu=neuer_wert,
            zustand=Dateizustand.von(pfad), backup=backup,
            ws=ws, eintrag=eintrag,
        )
