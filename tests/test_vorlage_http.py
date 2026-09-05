"""Die ausgelieferte Vorlage durch den echten HTTP-Weg - das kommt der Auslieferung
am nächsten.

Alle übrigen Tests laufen gegen das synthetische Workbook aus
``bestand.core.testing`` (siehe ``tests/conftest.py``). Diese Datei prüft
stattdessen die Datei, die tatsächlich mit ``START.bat``/``START.sh``
ausgeliefert wird: ``vorlage/Bestand- und Nachbestellungsliste 2026.xlsx``, mit
denselben Blättern, Merges, Formeln und Formatierungen wie die echte Mappe,
aber ohne Arbeitsdaten (siehe ``tests/test_template.py`` für die Struktur der
Vorlage selbst).

Die eingecheckte Datei wird dabei **nie** beschrieben - jeder Test arbeitet auf
einer Kopie in ``tmp_path``. Das wird am Ende zusätzlich über einen Hash der
Originaldatei nachgewiesen.
"""
from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import create_app
from app.settings import Einstellungen
from conftest import TEST_BASIS_URL

_WURZEL = Path(__file__).resolve().parent.parent
_VORLAGE = _WURZEL / "vorlage" / "Bestand- und Nachbestellungsliste 2026.xlsx"


def _blatt_raster() -> str:
    """Der Blattname der echten Mappe - aus ``config.json``, nicht hart verdrahtet."""
    with open(_WURZEL / "config.json", encoding="utf-8") as handle:
        return json.load(handle)["blatt_raster"]


def _hash(pfad: Path) -> str:
    return hashlib.sha256(pfad.read_bytes()).hexdigest()


def _formelzellen(pfad: Path, blatt: str) -> list[str]:
    """Koordinaten aller Zellen im Blatt, deren Wert eine Formel ist.

    ``data_only=False`` entspricht dem Ladepfad der Anwendung
    (``app.excel.lade_mappe``): eine gespeicherte Formel bleibt als Text, der
    mit ``=`` beginnt, erhalten - wäre sie stattdessen durch ihren letzten
    berechneten Wert ersetzt worden, stünde hier eine Zahl oder ``None``.
    """
    wb = load_workbook(str(pfad), data_only=False)
    ws = wb[blatt]
    return [
        cell.coordinate
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


@pytest.fixture()
def original_hash() -> str:
    """Hash der eingecheckten Vorlage vor dem Testlauf - Referenzwert für die Endkontrolle."""
    return _hash(_VORLAGE)


@pytest.fixture()
def vorlage_kopie(tmp_path: Path) -> Path:
    """Eine Arbeitskopie der Vorlage. Die eingecheckte Datei wird nie beschrieben."""
    ziel = tmp_path / _VORLAGE.name
    shutil.copyfile(_VORLAGE, ziel)
    return ziel


@pytest.fixture()
def einstellungen_vorlage(vorlage_kopie: Path) -> Einstellungen:
    return Einstellungen(
        iserv_domain="beispiel-schule.de",
        excel_pfad_kandidaten=(vorlage_kopie,),
        blatt_raster=_blatt_raster(),
    )


@pytest.fixture()
def client_vorlage(einstellungen_vorlage: Einstellungen) -> TestClient:
    """Eine isolierte App direkt auf der kopierten Vorlage, ohne echten IServ-Client."""
    application = create_app(einstellungen=einstellungen_vorlage)
    with TestClient(application, base_url=TEST_BASIS_URL) as testclient:
        yield testclient


# ── Lesen ─────────────────────────────────────────────────────────────────────

def test_health(client_vorlage: TestClient):
    antwort = client_vorlage.get("/health")
    assert antwort.status_code == 200
    assert antwort.json() == {"status": "ok"}


def test_startseite_rendert_html(client_vorlage: TestClient):
    antwort = client_vorlage.get("/")
    assert antwort.status_code == 200
    assert "text/html" in antwort.headers["content-type"]


def test_rows_struktur_der_leeren_vorlage(client_vorlage: TestClient):
    """Die Vorlage hat keine Arbeitsdaten - geprüft wird deshalb die Struktur, nicht ein Wert."""
    antwort = client_vorlage.get("/api/rows")
    assert antwort.status_code == 200
    daten = antwort.json()

    assert isinstance(daten["mtime"], (int, float))
    assert daten["cache_leer"] is True
    assert daten["zeilen"], "die Vorlage muss ein lesbares Raster liefern"
    for zeile in daten["zeilen"]:
        assert isinstance(zeile["key"], str) and zeile["key"]
        assert isinstance(zeile["fach"], str) and zeile["fach"]
        assert isinstance(zeile["jahrgang"], str) and zeile["jahrgang"]
        assert isinstance(zeile["bestand_ref"], str) and zeile["bestand_ref"]
        # Die Vorlage hat keine Arbeitsdaten - kein Wert ist vorbelegt.
        assert zeile["angemeldet"] is None
        assert zeile["bestand"] is None
        assert zeile["bestellt"] is None


# ── Schreiben ─────────────────────────────────────────────────────────────────

def test_schreiben_landet_wirklich_in_der_datei(
    client_vorlage: TestClient, vorlage_kopie: Path, einstellungen_vorlage: Einstellungen,
):
    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]

    antwort = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 42, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200, antwort.text

    nachher = client_vorlage.get("/api/rows").json()
    nachher_zeile = next(z for z in nachher["zeilen"] if z["key"] == zeile["key"])
    assert nachher_zeile["bestand"] == 42

    wb = load_workbook(str(vorlage_kopie))
    assert wb[einstellungen_vorlage.blatt_raster][zeile["bestand_ref"]].value == 42


def test_formeln_ueberleben_den_schreibvorgang(
    client_vorlage: TestClient, vorlage_kopie: Path, einstellungen_vorlage: Einstellungen,
):
    """Die wichtigste Regression: mit ``data_only=True`` gespeichert wären sie unwiderruflich weg."""
    blatt = einstellungen_vorlage.blatt_raster
    formeln_vorher = _formelzellen(vorlage_kopie, blatt)
    assert formeln_vorher, "die Vorlage muss tatsächlich Formeln enthalten, sonst prüft das nichts"

    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]
    antwort = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 7, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200, antwort.text

    formeln_nachher = _formelzellen(vorlage_kopie, blatt)
    assert formeln_nachher == formeln_vorher


def test_konflikt_bei_veralteter_mtime(client_vorlage: TestClient):
    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]
    antwort = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 3, "mtime": daten["mtime"] - 60,
    })
    assert antwort.status_code == 409


@pytest.mark.parametrize("leerer_wert", ["", None])
def test_leeren_wert_schreiben_loescht_die_zelle(client_vorlage: TestClient, leerer_wert):
    """Leer heißt 'nichts bestellt', 0 heißt 'nachgesehen, nichts offen' - die Mappe unterscheidet das."""
    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]
    erst = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 5, "mtime": daten["mtime"],
    })
    assert erst.status_code == 200, erst.text

    zweit = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": leerer_wert, "mtime": erst.json()["mtime"],
    })
    assert zweit.status_code == 200, zweit.text
    assert zweit.json()["zeile"]["bestand"] is None


def test_backup_wird_beim_schreiben_ueber_http_angelegt(client_vorlage: TestClient, vorlage_kopie: Path):
    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]
    antwort = client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 9, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200, antwort.text
    ordner = vorlage_kopie.parent / "backups"
    assert len(list(ordner.glob("*.xlsx"))) == 1


def test_original_vorlage_bleibt_unangetastet(client_vorlage: TestClient, original_hash: str):
    """Der wichtigste Nachweis dieser Datei: nie die eingecheckte Datei beschreiben."""
    daten = client_vorlage.get("/api/rows").json()
    zeile = daten["zeilen"][0]
    client_vorlage.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestand", "wert": 1, "mtime": daten["mtime"],
    })
    assert _hash(_VORLAGE) == original_hash
