"""Lesepfad über HTTP: Tabelle, JSON, fehlende Datei."""
from __future__ import annotations

import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import create_app
from app.settings import Einstellungen
from conftest import TEST_BASIS_URL


def _einrichtungs_app(tmp_path, einstellungen: Einstellungen) -> tuple[TestClient, Path]:
    config_pfad = tmp_path / "config.json"
    config_pfad.write_text(json.dumps({
        "iserv_domain": einstellungen.iserv_domain,
        "excel_pfad_kandidaten": [str(p) for p in einstellungen.excel_pfad_kandidaten],
        "blatt_raster": einstellungen.blatt_raster,
    }), encoding="utf-8")
    return (
        TestClient(create_app(einstellungen=einstellungen, config_pfad=config_pfad),
                   base_url=TEST_BASIS_URL),
        config_pfad,
    )


def test_startseite_zeigt_die_tabelle(client):
    antwort = client.get("/")
    assert antwort.status_code == 200
    text = antwort.text
    assert "Bestand und Nachbestellung" in text
    assert "Erdkunde" in text
    assert "5-6" in text          # Mehrjahresband
    assert "Latein" in text


def test_startseite_weist_auf_den_fehlenden_abruf_hin(client):
    """Ohne Cache stehen Titel und ISBN nicht zur Verfügung - das muss dastehen."""
    text = client.get("/").text
    assert "nach dem\nersten Abruf aus IServ" in text
    # Titel und ISBN kommen aus dem Cache: ohne Abruf steht das Leerzeichen da,
    # und der Sortierschlüssel derselben Zelle ist leer statt "—".
    assert text.count('data-wert="">—</td>') > 0


def test_api_rows_liefert_die_zeilen(client):
    daten = client.get("/api/rows").json()
    assert daten["cache_leer"] is True
    band = next(z for z in daten["zeilen"] if z["bestand_ref"] == "G3")
    assert (band["jahrgang"], band["angemeldet"], band["zu_bestellen"]) == ("5-6", 92, 2)


def test_fehlende_excel_datei_zeigt_die_geprueften_pfade(client, tmp_path):
    client.app.state.einstellungen = Einstellungen(
        iserv_domain="beispiel-schule.de",
        excel_pfad_kandidaten=(tmp_path / "a.xlsx", tmp_path / "b.xlsx"),
        blatt_raster="Bestand- und Nachbestellung",
    )
    antwort = client.get("/")
    assert antwort.status_code == 503
    assert "a.xlsx" in antwort.text and "b.xlsx" in antwort.text

    json_antwort = client.get("/api/rows")
    assert json_antwort.status_code == 503
    assert len(json_antwort.json()["geprueft"]) == 2


def test_fehlendes_blatt_meldet_klartext(client, einstellungen, workbook_path):
    client.app.state.einstellungen = Einstellungen(
        iserv_domain=einstellungen.iserv_domain,
        excel_pfad_kandidaten=(workbook_path,),
        blatt_raster="Gibt-es-nicht",
    )
    antwort = client.get("/")
    assert antwort.status_code == 500
    assert "Gibt-es-nicht" in antwort.text


def test_nur_bestellt_ist_ein_eingabefeld(client):
    """Die einzige Spalte, die von Hand geändert werden darf.

    "Bestand" hatte bis 2026-09-05 ebenfalls ein Feld. Es kommt aber aus IServ
    und wird bei jedem Abruf überschrieben - eine Eingabe dort hielt höchstens
    bis zum nächsten Abruf und sah bis dahin aus wie eine verlässliche Zahl.
    """
    text = client.get("/").text
    assert text.count('data-spalte="bestellt"') > 0
    assert 'data-spalte="bestand"' not in text
    assert 'data-spalte="angemeldet"' not in text
    assert 'data-spalte="zu_bestellen"' not in text


def test_die_seite_kennt_die_aenderungszeit(client):
    """Ohne mtime im HTML könnte der Browser keinen Konflikt erkennen."""
    daten = client.get("/api/rows").json()
    assert f'data-mtime="{daten["mtime"]}"' in client.get("/").text


def test_abruf_dialog_warnt_vor_dem_ueberschreiben(client):
    text = client.get("/").text
    assert "Aus IServ abrufen" in text
    assert "überschrieben" in text
    assert 'type="password"' in text


def test_einrichtung_prueft_die_mappe_vor_dem_speichern(tmp_path, einstellungen):
    unlesbar = tmp_path / "keine-echte-exceldatei.xlsx"
    unlesbar.write_text("keine Excel-Datei", encoding="utf-8")
    testclient, config_pfad = _einrichtungs_app(tmp_path, einstellungen)
    vorher = config_pfad.read_text(encoding="utf-8")
    with testclient:
        antwort = testclient.post("/api/einrichtung", json={"pfad": str(unlesbar)})
    assert antwort.status_code == 400
    assert "lesbare Excel" in antwort.json()["fehler"]
    assert config_pfad.read_text(encoding="utf-8") == vorher


def test_einrichtung_braucht_alle_dashboard_blaetter(tmp_path, einstellungen, leeres_workbook):
    unvollstaendig = tmp_path / "unvollstaendig.xlsx"
    wb = load_workbook(leeres_workbook)
    del wb["bestellt"]
    wb.save(unvollstaendig)
    testclient, config_pfad = _einrichtungs_app(tmp_path, einstellungen)
    vorher = config_pfad.read_text(encoding="utf-8")
    with testclient:
        antwort = testclient.post("/api/einrichtung", json={"pfad": str(unvollstaendig)})
    assert antwort.status_code == 400
    assert "bestellt" in antwort.json()["fehler"]
    assert config_pfad.read_text(encoding="utf-8") == vorher


def test_einrichtung_lehnt_leeres_raster_ab(tmp_path, einstellungen):
    leer = tmp_path / "leeres-raster.xlsx"
    wb = Workbook()
    wb.active.title = einstellungen.blatt_raster
    wb.create_sheet("bestellt")
    wb.create_sheet("zu Bestellen")
    wb.save(leer)
    testclient, config_pfad = _einrichtungs_app(tmp_path, einstellungen)
    vorher = config_pfad.read_text(encoding="utf-8")

    with testclient:
        antwort = testclient.post("/api/einrichtung", json={"pfad": str(leer)})

    assert antwort.status_code == 400
    assert "Bestandsraster" in antwort.json()["fehler"]
    assert config_pfad.read_text(encoding="utf-8") == vorher


def test_einrichtung_speichert_eine_gueltige_mappe(
    tmp_path, einstellungen, leeres_workbook,
):
    testclient, config_pfad = _einrichtungs_app(tmp_path, einstellungen)

    with testclient:
        antwort = testclient.post("/api/einrichtung", json={"pfad": str(leeres_workbook)})

    assert antwort.status_code == 200
    assert antwort.json() == {"ok": True}
    gespeichert = json.loads(config_pfad.read_text(encoding="utf-8"))
    assert gespeichert["excel_pfad_kandidaten"][0] == str(leeres_workbook)
