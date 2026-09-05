"""POST /api/cell: nur Bestellt, nur über den Zeilenschlüssel."""
from __future__ import annotations

from pathlib import Path
from threading import Event, Thread

import pytest
from bestand.core import parse_grid
from bestand.core.testing import SHEET_NAME
from openpyxl import load_workbook

from app import excel as excel_modul
from app.excel import (
    Dateizustand,
    Konflikt,
    UngueltigeAenderung,
    kuerze_backups,
    pruefe_wert,
    schreibe_zelle,
)


def _zeilen(client):
    antwort = client.get("/api/rows")
    assert antwort.status_code == 200
    return antwort.json()


def _zeile(daten, bestand_ref):
    return next(z for z in daten["zeilen"] if z["bestand_ref"] == bestand_ref)


def test_bestellt_wird_geschrieben_und_zurueckgemeldet(client, workbook_path: Path):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")          # Erdkunde Jg 5-6, Bestand 60, Bestellt 30
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 71, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200, antwort.text
    körper = antwort.json()
    assert körper["ref"] == "H3"         # die Bestellt-Zelle desselben Bandes
    assert körper["zeile"]["bestellt"] == 71
    # 92 angemeldet - 60 Bestand - 71 bestellt = -39: der Bedarf wird sofort neu gerechnet.
    assert körper["zeile"]["zu_bestellen"] == -39
    assert körper["mtime"] != daten["mtime"]

    wb = load_workbook(str(workbook_path))
    assert wb[SHEET_NAME]["H3"].value == 71


def test_leerer_wert_loescht_die_zelle(client):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": "", "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200
    assert antwort.json()["zeile"]["bestellt"] is None


def test_formelspalte_bleibt_nach_dem_schreiben_formel(client, workbook_path: Path):
    """Der Schreibpfad lädt mit data_only=False - sonst wären die Formeln danach weg."""
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 5, "mtime": daten["mtime"],
    })
    wb = load_workbook(str(workbook_path))
    assert wb[SHEET_NAME]["I3"].value == "=F3+F4-G3-H3"


@pytest.mark.parametrize(
    "spalte", ["bestand", "angemeldet", "zu_bestellen", "Bestellt", "", "stand"]
)
def test_andere_spalten_werden_abgelehnt(client, spalte):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": spalte, "wert": 1, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 400
    assert "fehler" in antwort.json()


def test_unbekannter_schluessel_wird_abgelehnt(client):
    daten = _zeilen(client)
    antwort = client.post("/api/cell", json={
        "key": "0:Gibt-Es-Nicht:ZZ99", "spalte": "bestellt", "wert": 1,
        "mtime": daten["mtime"],
    })
    assert antwort.status_code == 400
    assert "neu laden" in antwort.json()["fehler"]


def test_zellbezug_statt_schluessel_schreibt_nichts(client, workbook_path: Path):
    """Ein freier Bezug ist kein Schlüssel - die Route löst ihn nicht auf."""
    vorher = load_workbook(str(workbook_path))[SHEET_NAME]["C3"].value
    daten = _zeilen(client)
    antwort = client.post("/api/cell", json={
        "key": "C3", "spalte": "bestellt", "wert": 999, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 400
    assert load_workbook(str(workbook_path))[SHEET_NAME]["C3"].value == vorher


@pytest.mark.parametrize("wert", [-1, 3.5, "viele", True, [], {"a": 1}])
def test_ungueltige_werte_werden_abgelehnt(client, wert):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": wert, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 400


def test_veraltete_mtime_ist_ein_konflikt(client, workbook_path: Path):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 7, "mtime": daten["mtime"] - 60,
    })
    assert antwort.status_code == 409
    körper = antwort.json()
    assert "neu laden" in körper["fehler"]
    assert körper["mtime"] == pytest.approx(daten["mtime"])
    # Nichts geschrieben.
    assert load_workbook(str(workbook_path))[SHEET_NAME]["H3"].value == 30


def test_zweiter_schreibvorgang_mit_alter_mtime_scheitert(client):
    """Genau der Fall, für den das optimistische Sperren da ist."""
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    erst = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 1, "mtime": daten["mtime"],
    })
    assert erst.status_code == 200
    zweit = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 2, "mtime": daten["mtime"],
    })
    assert zweit.status_code == 409
    # Mit der neuen mtime aus der ersten Antwort geht es weiter.
    dritt = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 2, "mtime": erst.json()["mtime"],
    })
    assert dritt.status_code == 200


def test_ohne_mtime_wird_abgelehnt(client):
    """Ohne gesehenen Versionsstand darf eine API-Anfrage nicht schreiben."""
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 3,
    })
    assert antwort.status_code == 400
    assert "Änderungszeit" in antwort.json()["fehler"]


def test_backup_wird_angelegt(client, workbook_path: Path):
    daten = _zeilen(client)
    zeile = _zeile(daten, "G3")
    antwort = client.post("/api/cell", json={
        "key": zeile["key"], "spalte": "bestellt", "wert": 4, "mtime": daten["mtime"],
    })
    assert antwort.status_code == 200
    name = antwort.json()["backup"]
    assert name is not None
    assert (workbook_path.parent / "backups" / name).is_file()


def test_backups_werden_auf_die_eingestellte_zahl_gekuerzt(client, einstellungen,
                                                           workbook_path: Path):
    """backups_behalten=3: nach fünf Änderungen liegen drei Dateien im Ordner."""
    object.__setattr__(einstellungen, "backups_behalten", 3)
    mtime = _zeilen(client)["mtime"]
    key = _zeile(_zeilen(client), "G3")["key"]
    for wert in range(5):
        antwort = client.post("/api/cell", json={
            "key": key, "spalte": "bestellt", "wert": wert, "mtime": mtime,
        })
        assert antwort.status_code == 200, antwort.text
        mtime = antwort.json()["mtime"]
    ordner = workbook_path.parent / "backups"
    assert len(list(ordner.glob("*.xlsx"))) == 3


def test_das_raster_wird_bei_jedem_schreiben_neu_geparst(client, workbook_path: Path):
    """Der Schlüssel zeigt auf eine Zeile, nicht auf eine Zellkoordinate."""
    wb = load_workbook(str(workbook_path))
    grid = parse_grid(wb[SHEET_NAME])
    schluessel = {e.key for e in grid.entries}
    daten = _zeilen(client)
    assert {z["key"] for z in daten["zeilen"]} == schluessel


def test_gleichzeitige_schreibvorgaenge_koennen_keine_aenderung_ueberschreiben(
    workbook_path: Path, monkeypatch,
):
    """Der zweite Schreiber wartet, prüft danach seine alte Version und scheitert.

    Der erste Save wird kontrolliert angehalten. Ohne das gemeinsame Schloss
    würden inzwischen beide Threads ihre jeweils alte Workbook-Kopie speichern;
    die zweite Speicherung würde die erste still verlieren lassen.
    """
    wb = load_workbook(str(workbook_path))
    entry = parse_grid(wb[SHEET_NAME]).entries[0]
    initial_mtime = Dateizustand.von(workbook_path).mtime
    erster_save_begonnen = Event()
    save_freigeben = Event()
    zweiter_gestartet = Event()
    zweiter_fertig = Event()
    original_save = excel_modul.atomic_save_workbook

    def angehaltener_save(*args, **kwargs):
        erster_save_begonnen.set()
        assert save_freigeben.wait(timeout=5)
        return original_save(*args, **kwargs)

    monkeypatch.setattr(excel_modul, "atomic_save_workbook", angehaltener_save)
    ergebnisse: dict[str, object] = {}

    def erster_schreiber():
        try:
            ergebnisse["erster"] = schreibe_zelle(
                workbook_path, SHEET_NAME, key=entry.key, spalte="bestellt", wert=71,
                mtime=initial_mtime,
            )
        except BaseException as exc:  # noqa: BLE001 - Assertion für den Thread sammeln
            ergebnisse["erster"] = exc

    def zweiter_schreiber():
        zweiter_gestartet.set()
        try:
            ergebnisse["zweiter"] = schreibe_zelle(
                workbook_path, SHEET_NAME, key=entry.key, spalte="bestellt", wert=72,
                mtime=initial_mtime,
            )
        except BaseException as exc:  # noqa: BLE001 - Assertion für den Thread sammeln
            ergebnisse["zweiter"] = exc
        finally:
            zweiter_fertig.set()

    thread_1 = Thread(target=erster_schreiber)
    thread_1.start()
    assert erster_save_begonnen.wait(timeout=5)
    thread_2 = Thread(target=zweiter_schreiber)
    thread_2.start()
    assert zweiter_gestartet.wait(timeout=5)
    assert not zweiter_fertig.wait(timeout=0.2)
    save_freigeben.set()
    thread_1.join(timeout=5)
    thread_2.join(timeout=5)

    assert not thread_1.is_alive()
    assert not thread_2.is_alive()
    assert not isinstance(ergebnisse["erster"], BaseException)
    assert isinstance(ergebnisse["zweiter"], Konflikt)
    assert load_workbook(str(workbook_path))[SHEET_NAME][entry.slots["bestellt"].ref].value == 71


# ── Einheiten ohne HTTP ───────────────────────────────────────────────────────

@pytest.mark.parametrize("roh,erwartet", [
    (None, None), ("", None), ("  ", None), (0, 0), (12, 12), ("12", 12), (7.0, 7),
])
def test_pruefe_wert_akzeptiert(roh, erwartet):
    assert pruefe_wert(roh) == erwartet


@pytest.mark.parametrize("roh", [-1, "-1", 2.5, "zwei", True, False, [], object()])
def test_pruefe_wert_lehnt_ab(roh):
    with pytest.raises(UngueltigeAenderung):
        pruefe_wert(roh)


def test_kuerze_backups_behaelt_die_neuesten(tmp_path: Path):
    ordner = tmp_path / "backups"
    ordner.mkdir()
    import os
    for i in range(6):
        pfad = ordner / f"Bestand.2026090{i}-120000.xlsx"
        pfad.write_bytes(b"x")
        os.utime(pfad, (1_000_000 + i, 1_000_000 + i))
    geloescht = kuerze_backups(ordner, 2)
    verblieben = sorted(p.name for p in ordner.glob("*.xlsx"))
    assert len(geloescht) == 4
    assert verblieben == ["Bestand.20260904-120000.xlsx", "Bestand.20260905-120000.xlsx"]


def test_kuerze_backups_ohne_ordner_ist_kein_fehler(tmp_path: Path):
    assert kuerze_backups(tmp_path / "gibt-es-nicht", 5) == []
