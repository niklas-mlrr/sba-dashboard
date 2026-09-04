"""Die veröffentlichte Excel-Vorlage muss das Dashboard ohne echte Daten tragen."""
from __future__ import annotations

from pathlib import Path

from bestand.core import parse_grid
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

VORLAGE = (
    Path(__file__).parent.parent / "vorlage" / "Bestand- und Nachbestellungsliste 2026.xlsx"
)


def test_vorlage_hat_das_erwartete_raster_ohne_veraenderliche_zahlen():
    wb = load_workbook(VORLAGE, data_only=False)
    ws = wb["Bestand- und Nachbestellung"]
    grid = parse_grid(ws)

    assert wb.sheetnames == ["Bestand- und Nachbestellung", "zu Bestellen", "bestellt", "erhalten"]
    assert len(grid.entries) == 72
    assert sum(cell.data_type == "f" for row in ws.iter_rows() for cell in row) == 72

    for eintrag in grid.entries:
        for ref in (*eintrag.angemeldet_refs, eintrag.slots["bestand"].ref,
                    eintrag.slots["bestellt"].ref):
            assert ws[ref].value is None


def test_vorlage_enthaelt_keine_eingegebenen_tabellenwerte():
    wb = load_workbook(VORLAGE, data_only=False)
    for ws in wb.worksheets:
        for tabelle in ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(tabelle.ref)
            for zeile in ws.iter_rows(
                min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                for zelle in zeile:
                    assert zelle.value is None or zelle.data_type == "f"
