"""Regressionstests für die Nebenläufigkeitsgarantien von ``app/excel.py``.

``tests/test_lock.py`` deckt bereits die prozessübergreifende Nachbardatei-
Sperre ab (zwei getrennte Prozesse), ``tests/test_cell_write.py`` bereits den
kontrolliert angehaltenen Save zweier Threads, die um dieselbe Zelle
konkurrieren. Diese Datei dupliziert das nicht, sondern ergänzt:

* eine echte Race ohne künstlich verlangsamten Save (der Sieger entscheidet
  sich rein über die Reihenfolge des Schloss-Erwerbs, nicht über einen Monkey-
  Patch),
* viele Schreiber mit Wiederholung nach ``Konflikt``,
* dass ``_prozesssperre`` denselben Pfad unabhängig vom Arbeitsverzeichnis
  (und, wo möglich, über einen Symlink) auf dasselbe Schloss abbildet,
* ``wartezeit=0`` auf eine bereits gehaltene Sperre,
* dass ein Lesevorgang während eines Schreibvorgangs nie eine halbe Mappe
  sieht,
* dass Refresh und manuelle Zelländerung dasselbe Schloss benutzen.

Unter Windows liest ``_versuche_dateisperre`` mit ``msvcrt.locking`` statt
``fcntl.flock`` - deshalb wird hier ausschließlich über die plattform-
unabhängige Schnittstelle (``arbeitsmappe_sperren``, ``schreibe_zelle``,
``RefreshManager``) geprüft, nie über ``chmod``, ``os.getuid`` oder sonstige
POSIX-Annahmen. Wo ein Fall unter Windows sinnlos ist (Symlinks brauchen dort
Admin-Rechte oder den Entwicklermodus), ist er mit ``skipif`` markiert.
"""
from __future__ import annotations

import os
import time
from pathlib import Path
from threading import Barrier, Event, Thread

import pytest
from bestand.core import parse_grid
from bestand.core.testing import SHEET_NAME, FakeClient
from openpyxl import load_workbook

from app import excel as excel_modul
from app import refresh as refresh_modul
from app.excel import (
    Dateizustand,
    Gesperrt,
    Konflikt,
    Schreibergebnis,
    arbeitsmappe_sperren,
    schreibe_zelle,
)
from app.refresh import RefreshManager
from app.settings import Einstellungen

# Ein Thread-Schreibvorgang (Excel laden, ändern, atomar speichern) dauert auf
# gewöhnlicher Hardware deutlich unter einer Sekunde; 20 Versuche je Thread und
# Timeouts von 5-10s je Wartepunkt reichen mit reichlich Marge und halten die
# gesamte Datei unter etwa 10s.
_TIMEOUT = 5.0


# ── Echte Race, kein kontrollierter Save ─────────────────────────────────────

def test_zwei_gleichzeitige_schreiber_auf_dieselbe_mtime_nur_einer_gewinnt(workbook_path: Path):
    """Ohne jeden Monkey-Patch: der Sieger entscheidet sich am Schloss, nicht am Timing.

    Anders als ``test_cell_write.py::test_gleichzeitige_schreibvorgaenge_...``
    wird der Save hier nicht künstlich angehalten - die Invariante ("genau
    einer gewinnt") muss auch bei echtem, unverlangsamtem Wettlauf gelten.
    """
    wb = load_workbook(str(workbook_path))
    entry = parse_grid(wb[SHEET_NAME]).entries[0]
    start_mtime = Dateizustand.von(workbook_path).mtime
    start = Barrier(2, timeout=_TIMEOUT)
    ergebnisse: dict[str, object] = {}

    def schreiber(name: str, wert: int) -> None:
        start.wait()
        try:
            ergebnisse[name] = schreibe_zelle(
                workbook_path, SHEET_NAME, key=entry.key, spalte="bestellt", wert=wert,
                mtime=start_mtime,
            )
        except BaseException as exc:  # noqa: BLE001 - Ergebnis für die Haupt-Assertion sammeln
            ergebnisse[name] = exc

    t1 = Thread(target=schreiber, args=("a", 21))
    t2 = Thread(target=schreiber, args=("b", 22))
    t1.start()
    t2.start()
    t1.join(timeout=_TIMEOUT)
    t2.join(timeout=_TIMEOUT)
    assert not t1.is_alive() and not t2.is_alive()

    erfolge = [n for n, e in ergebnisse.items() if isinstance(e, Schreibergebnis)]
    konflikte = [n for n, e in ergebnisse.items() if isinstance(e, Konflikt)]
    assert len(erfolge) == 1, ergebnisse
    assert len(konflikte) == 1, ergebnisse

    gewinner: Schreibergebnis = ergebnisse[erfolge[0]]  # type: ignore[assignment]
    wb_final = load_workbook(str(workbook_path))
    assert wb_final[SHEET_NAME][entry.slots["bestellt"].ref].value == gewinner.neu


def test_acht_threads_schreiben_mit_wiederholung_nach_konflikt(workbook_path: Path):
    """Viele Schreiber, jeder mit der zuletzt gesehenen ``mtime``, mit Retry nach Konflikt.

    Am Ende: jeder Schreibvorgang ist entweder durchgekommen oder wurde
    abgelehnt (nie eine stille Doppel-Anwendung), und die Mappe lässt sich
    danach noch laden.
    """
    wb = load_workbook(str(workbook_path))
    # Eine andere Zeile als die vorherige Race, damit die beiden Tests sich
    # nicht gegenseitig beeinflussen können, liefen sie je aus Versehen parallel.
    entry = parse_grid(wb[SHEET_NAME]).entries[1]
    anzahl = 8
    ergebnisse: dict[int, object] = {}

    def schreiber(index: int) -> None:
        wert = 100 + index
        for _ in range(20):
            mtime = Dateizustand.von(workbook_path).mtime
            try:
                ergebnisse[index] = schreibe_zelle(
                    workbook_path, SHEET_NAME, key=entry.key, spalte="bestellt", wert=wert,
                    mtime=mtime,
                )
                return
            except Konflikt:
                continue
        ergebnisse[index] = TimeoutError(f"Thread {index} kam nach 20 Versuchen nicht durch.")

    threads = [Thread(target=schreiber, args=(i,)) for i in range(anzahl)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=10)
    assert all(not t.is_alive() for t in threads)

    for index in range(anzahl):
        ergebnis = ergebnisse.get(index)
        assert isinstance(ergebnis, Schreibergebnis), f"Thread {index}: {ergebnis!r}"

    # Die Mappe ist danach intakt und ladbar - keine halbgeschriebene Datei.
    wb_final = load_workbook(str(workbook_path))
    endwert = wb_final[SHEET_NAME][entry.slots["bestellt"].ref].value
    assert endwert in {100 + i for i in range(anzahl)}


# ── Verzeichnisunabhängigkeit der Prozesssperre ──────────────────────────────

def test_prozesssperre_ist_unabhaengig_vom_arbeitsverzeichnis(tmp_path: Path, monkeypatch):
    """``_prozesssperre`` löst über ``Path.resolve()`` auf - Groß-/Kleinschreibung des
    Arbeitsverzeichnisses oder ein relativer Bezug dürfen kein zweites Schloss ergeben.
    """
    datei = tmp_path / "Mappe.xlsx"
    datei.write_bytes(b"x")
    monkeypatch.chdir(tmp_path)

    absolut = datei
    relativ = Path("Mappe.xlsx")
    assert not relativ.is_absolute()
    assert excel_modul._prozesssperre(absolut) is excel_modul._prozesssperre(relativ)


@pytest.mark.skipif(
    os.name == "nt",
    reason="Symlinks brauchen unter Windows Admin-Rechte oder den Entwicklermodus.",
)
def test_prozesssperre_ist_unabhaengig_vom_verweis_ueber_einen_symlink(tmp_path: Path):
    datei = tmp_path / "Mappe.xlsx"
    datei.write_bytes(b"x")
    verweis = tmp_path / "Verweis.xlsx"
    verweis.symlink_to(datei)

    assert excel_modul._prozesssperre(datei) is excel_modul._prozesssperre(verweis)


# ── wartezeit=0 ───────────────────────────────────────────────────────────────

def test_wartezeit_null_scheitert_sofort_und_stoert_den_ersten_halter_nicht(workbook_path: Path):
    haelt = Event()
    freigeben = Event()
    ergebnisse: dict[str, str] = {}

    def erster_halter() -> None:
        with arbeitsmappe_sperren(workbook_path, wartezeit=_TIMEOUT):
            haelt.set()
            assert freigeben.wait(timeout=_TIMEOUT)
        ergebnisse["erster"] = "fertig"

    thread = Thread(target=erster_halter)
    thread.start()
    assert haelt.wait(timeout=_TIMEOUT)

    with pytest.raises(Gesperrt):
        with arbeitsmappe_sperren(workbook_path, wartezeit=0):
            pass  # pragma: no cover - darf nie erreicht werden

    freigeben.set()
    thread.join(timeout=_TIMEOUT)
    assert not thread.is_alive()
    assert ergebnisse.get("erster") == "fertig", (
        "die erste Sperre muss trotz des Konkurrenten sauber fertig werden"
    )


def test_wartezeit_negativ_wird_abgelehnt(workbook_path: Path):
    with pytest.raises(ValueError):
        with arbeitsmappe_sperren(workbook_path, wartezeit=-1):
            pass  # pragma: no cover


# ── Lesen während eines Schreibvorgangs ──────────────────────────────────────

def test_lesen_waehrend_des_schreibens_liefert_nie_eine_halbe_mappe(
    client, workbook_path: Path, monkeypatch,
):
    """``GET /api/rows`` während eines angehaltenen Saves: nie der alte *und* der neue Stand
    vermischt, immer eine vollständig ladbare, konsistente Antwort.

    Muster wie ``test_cell_write.py::test_gleichzeitige_schreibvorgaenge_...``:
    der Save wird kontrolliert angehalten, bevor er die Datei tatsächlich
    ersetzt - solange er dort hängt, ist die Datei auf der Platte noch
    vollständig die alte Fassung.
    """
    save_begonnen = Event()
    save_freigeben = Event()
    original_save = excel_modul.atomic_save_workbook

    def angehaltener_save(*args, **kwargs):
        save_begonnen.set()
        assert save_freigeben.wait(timeout=_TIMEOUT)
        return original_save(*args, **kwargs)

    monkeypatch.setattr(excel_modul, "atomic_save_workbook", angehaltener_save)

    vorher = client.get("/api/rows").json()
    zeile = vorher["zeilen"][0]
    schreib_antwort: dict[str, object] = {}

    def schreiben() -> None:
        schreib_antwort["antwort"] = client.post("/api/cell", json={
            "key": zeile["key"], "spalte": "bestellt", "wert": 111, "mtime": vorher["mtime"],
        })

    thread = Thread(target=schreiben)
    thread.start()
    assert save_begonnen.wait(timeout=_TIMEOUT)

    # Während der Save hängt, ist auf der Platte garantiert noch die alte
    # Fassung - das Lesen darf hier niemals scheitern oder etwas Halbes liefern.
    waehrend_antwort = client.get("/api/rows")
    assert waehrend_antwort.status_code == 200
    waehrend = waehrend_antwort.json()
    zeile_waehrend = next(z for z in waehrend["zeilen"] if z["key"] == zeile["key"])
    assert zeile_waehrend["bestellt"] == zeile["bestellt"]

    save_freigeben.set()
    thread.join(timeout=_TIMEOUT)
    assert not thread.is_alive()
    assert schreib_antwort["antwort"].status_code == 200, schreib_antwort["antwort"].text

    nachher = client.get("/api/rows").json()
    zeile_nachher = next(z for z in nachher["zeilen"] if z["key"] == zeile["key"])
    assert zeile_nachher["bestellt"] == 111


# ── Refresh und Zelländerung teilen sich das Schloss ─────────────────────────

def test_refresh_und_zellaenderung_teilen_sich_das_schloss(
    einstellungen: Einstellungen, workbook_path: Path, monkeypatch,
):
    """``RefreshManager._aktualisiere_mappe`` und ``schreibe_zelle`` sperren über
    dieselbe Funktion (``arbeitsmappe_sperren``). Solange der Refresh die Sperre
    hält, muss ein gleichzeitiger manueller Schreibversuch mit kurzer Wartezeit
    ``Gesperrt`` bekommen, statt gleichzeitig zu schreiben.
    """
    haelt_lock = Event()
    freigeben = Event()
    original_apply_snapshot = refresh_modul.apply_snapshot

    def blockierendes_apply_snapshot(*args, **kwargs):
        # Wird innerhalb von "with arbeitsmappe_sperren(pfad):" aufgerufen -
        # ab hier hält der Refresh nachweislich die Sperre.
        haelt_lock.set()
        assert freigeben.wait(timeout=_TIMEOUT)
        return original_apply_snapshot(*args, **kwargs)

    monkeypatch.setattr(refresh_modul, "apply_snapshot", blockierendes_apply_snapshot)

    manager = RefreshManager()
    manager.starte(einstellungen, FakeClient())
    assert haelt_lock.wait(timeout=_TIMEOUT)

    with pytest.raises(Gesperrt):
        with arbeitsmappe_sperren(workbook_path, wartezeit=0.2):
            pass  # pragma: no cover - darf nie erreicht werden

    freigeben.set()
    frist = time.monotonic() + _TIMEOUT
    while time.monotonic() < frist and manager.laeuft():
        time.sleep(0.02)
    assert not manager.laeuft(), "der Refresh muss nach Freigabe zu Ende kommen"

    status = manager.status()
    assert status is not None and status["fehler"] is None, status

    # Nach dem Refresh lässt sich die Mappe weiterhin normal beschreiben - die
    # Sperre wurde sauber freigegeben, nicht liegen gelassen.
    daten_nachher = load_workbook(str(workbook_path))
    entry = parse_grid(daten_nachher[SHEET_NAME]).entries[0]
    ergebnis = schreibe_zelle(
        workbook_path, SHEET_NAME, key=entry.key, spalte="bestellt", wert=5,
        mtime=Dateizustand.von(workbook_path).mtime,
    )
    assert ergebnis.neu == 5
