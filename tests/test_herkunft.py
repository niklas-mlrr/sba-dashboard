"""Host- und Origin-Prüfung: was die Bindung an 127.0.0.1 allein offen lässt.

Die Begründung beider Schichten steht in ``app/sicherheit.py``. Hier steht der
Nachweis, dass sie greifen - und dass sie die Wege *nicht* zumachen, auf denen
das Dashboard tatsächlich benutzt wird.
"""
from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from app.main import create_app
from app.settings import Einstellungen
from app.sicherheit import herkunft_erlaubt
from conftest import TEST_BASIS_URL


@pytest.fixture()
def fremder_client(einstellungen: Einstellungen) -> TestClient:
    """Derselbe Server, aber unter einem fremden Namen angesprochen.

    Genau das tut DNS-Rebinding: die Adresse zeigt auf 127.0.0.1, der
    ``Host``-Kopf trägt aber weiter die Domain des Angreifers.
    """
    with TestClient(create_app(einstellungen=einstellungen),
                    base_url="http://boesewicht.example") as testclient:
        yield testclient


# ── Host ──────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("pfad", ["/", "/api/rows", "/health"])
def test_fremder_host_wird_abgewiesen(fremder_client: TestClient, pfad: str):
    """Auch die Lesewege - dort stehen die Anmeldezahlen je Jahrgang."""
    assert fremder_client.get(pfad).status_code == 400


@pytest.mark.parametrize("basis", [
    "http://127.0.0.1",
    "http://127.0.0.1:8765",   # der Port, den freier_port normalerweise nimmt
    "http://127.0.0.1:8770",   # und der, auf den ein zweites Fenster ausweicht
    "http://localhost:8765",
])
def test_eigene_hosts_kommen_durch(einstellungen: Einstellungen, basis: str):
    """Jeder Port muss ohne Konfiguration passen - freier_port weicht aus."""
    with TestClient(create_app(einstellungen=einstellungen), base_url=basis) as testclient:
        assert testclient.get("/health").status_code == 200


# ── Origin ────────────────────────────────────────────────────────────────────

def test_beenden_von_fremder_seite_wird_abgelehnt(client: TestClient):
    """Der Anlassfall: ein POST ohne Körper aus einem beliebigen Nachbartab."""
    class _Server:
        should_exit = False

    server = _Server()
    client.app.state.server = server
    try:
        antwort = client.post("/api/beenden", headers={"Origin": "https://boesewicht.example"})
        assert antwort.status_code == 403
        assert "fremden Internetseite" in antwort.json()["fehler"]
        # Der Punkt der Prüfung: die Wirkung bleibt aus, nicht nur die Antwort.
        assert server.should_exit is False
    finally:
        client.app.state.server = None


def test_beenden_mit_eigenem_origin_geht_durch(client: TestClient):
    class _Server:
        should_exit = False

    server = _Server()
    client.app.state.server = server
    try:
        antwort = client.post("/api/beenden", headers={"Origin": "http://127.0.0.1:8765"})
        assert antwort.status_code == 200
        assert server.should_exit is True
    finally:
        client.app.state.server = None


def test_schreiben_von_fremder_seite_aendert_die_mappe_nicht(client: TestClient, workbook_path):
    from bestand.core.testing import SHEET_NAME
    from openpyxl import load_workbook

    daten = client.get("/api/rows").json()
    zeile = next(z for z in daten["zeilen"] if z["bestand_ref"] == "G3")
    vorher = load_workbook(str(workbook_path))[SHEET_NAME]["G3"].value

    antwort = client.post(
        "/api/cell",
        json={"key": zeile["key"], "spalte": "bestellt", "wert": 999, "mtime": daten["mtime"]},
        headers={"Origin": "http://boesewicht.example"},
    )
    assert antwort.status_code == 403
    assert load_workbook(str(workbook_path))[SHEET_NAME]["G3"].value == vorher


def test_lesen_mit_fremdem_origin_bleibt_erlaubt(client: TestClient):
    """GET ändert nichts; gegen das Mitlesen schützt die Same-Origin-Policy.

    Gegen den einen Fall, in dem sie es nicht tut - DNS-Rebinding -, schützt
    die Host-Prüfung eine Schicht darüber. Ein Origin-Verbot auf GET brächte
    darüber hinaus nichts und bräche stattdessen jeden Aufruf ohne Browser.
    """
    assert client.get("/api/rows", headers={"Origin": "http://boesewicht.example"}).status_code == 200


def test_ohne_origin_bleibt_alles_wie_vorher(client: TestClient):
    """curl, tools/diagnose.py und der TestClient schicken keinen Origin."""
    assert client.get("/health").status_code == 200
    assert client.post("/api/refresh", json={}).status_code == 400  # 400, nicht 403


@pytest.mark.parametrize("origin,erlaubt", [
    (None, True),
    ("http://127.0.0.1:8765", True),
    ("http://localhost", True),
    ("https://127.0.0.1", True),
    ("http://boesewicht.example", False),
    ("http://127.0.0.1.boesewicht.example", False),   # Präfix, nicht Host
    ("http://xn--127-0-0-1", False),
    ("null", False),                                   # sandboxed iframe, file://
    ("", False),
])
def test_herkunft_erlaubt(origin: str | None, erlaubt: bool):
    assert herkunft_erlaubt(origin) is erlaubt


def test_der_testclient_spricht_die_produktive_adresse(client: TestClient):
    """Sonst prüfte die Suite eine Anwendung, die es so nicht gibt."""
    assert TEST_BASIS_URL == "http://127.0.0.1"
    assert client.get("/health").request.headers["host"] == "127.0.0.1"
