"""POST /api/refresh: ein Lauf, Fehler in Klartext, Passwort nirgends.

Der Abruf ist die einzige Stelle, an der das Dashboard Zugangsdaten anfasst.
Die Tests prüfen deshalb nicht nur, dass er funktioniert, sondern auch, dass
das Passwort keine der Antworten, keinen Log-Eintrag und keinen Modulzustand
erreicht.
"""
from __future__ import annotations

import json
import logging
import threading
import time
from dataclasses import fields
from pathlib import Path

import pytest
from bestand.core.testing import SHEET_NAME, FakeClient
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import cache as cache_modul
from app.main import create_app
from app.refresh import Lauf, RefreshManager
from conftest import TEST_BASIS_URL

PASSWORT = "geheim-Kennwort-2026!"
BENUTZER = "b.lehrer"


def _warte_auf_ende(client, sekunden: float = 10.0) -> dict:
    """Pollt /api/refresh/status, bis der Lauf fertig ist."""
    frist = time.monotonic() + sekunden
    while time.monotonic() < frist:
        stand = client.get("/api/refresh/status").json()
        if stand["fertig"]:
            return stand
        time.sleep(0.02)
    raise AssertionError(f"Abruf wurde nicht fertig: {stand}")


def _abrufen(client, factory=FakeClient, **felder):
    client.app.state.client_factory = factory
    nutzlast = {"benutzer": BENUTZER, "passwort": PASSWORT}
    nutzlast.update(felder)
    return client.post("/api/refresh", json=nutzlast)


def _manager(client) -> RefreshManager:
    return client.app.state.refresh_manager


def test_app_instanz_hat_eigenen_refresh_manager(einstellungen):
    erste_app = create_app(einstellungen=einstellungen)
    zweite_app = create_app(einstellungen=einstellungen)

    with TestClient(erste_app, base_url=TEST_BASIS_URL), \
            TestClient(zweite_app, base_url=TEST_BASIS_URL):
        assert erste_app.state.refresh_manager is not zweite_app.state.refresh_manager


class _FehlerClient:
    """Client-Fabrik, deren ``login()`` eine bestimmte Ausnahme wirft."""

    def __init__(self, ausnahme):
        self.ausnahme = ausnahme

    def __call__(self, *args, **kwargs):
        self.args = args
        return self

    def login(self):
        raise self.ausnahme


class _BlockierenderClient(FakeClient):
    """Hält den Lauf in der Serienabfrage an, bis ``freigabe`` gesetzt wird."""

    freigabe = threading.Event()

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        eltern = self.series

        class _Warten:
            def get_all(self_inner, detailed=False):
                _BlockierenderClient.freigabe.wait(5)
                return eltern.get_all(detailed=detailed)

        self.series = _Warten()


# ── Erfolgsfall ───────────────────────────────────────────────────────────────

def test_abruf_schreibt_zahlen_und_cache(client, workbook_path: Path):
    antwort = _abrufen(client)
    assert antwort.status_code == 202, antwort.text
    assert antwort.json()["job_id"]

    stand = _warte_auf_ende(client)
    assert stand["fehler"] is None, stand
    assert stand["fehlercode"] is None
    assert stand["fortschritt"] == 100
    assert stand["zusammenfassung"]["schuljahr"] == "2026/2027"
    assert stand["zusammenfassung"]["geaendert"] > 0

    wb = load_workbook(str(workbook_path))
    # Terra 5/6: 60 im Bestand laut Fake-IServ, in G3 (Jg 5-6 verbunden).
    assert wb[SHEET_NAME]["G3"].value == 60
    # Die Formelspalte hat den Abruf unbeschadet überstanden.
    assert wb[SHEET_NAME]["I3"].value == "=F3+F4-G3-H3"


def test_cache_bekommt_formatierte_isbn(client, workbook_path: Path):
    """Im Live-Test stand in der Mappe 978-3-…, im Cache die nackte Ziffernfolge."""
    _abrufen(client)
    _warte_auf_ende(client)

    cache = cache_modul.laden(workbook_path)
    assert not cache.leer
    assert cache.schuljahr == "2026/2027"
    isbns = [e.isbn for e in cache.eintraege.values() if e.isbn]
    assert isbns
    assert all("-" in isbn for isbn in isbns), isbns
    titel = [e.titel for e in cache.eintraege.values() if e.titel]
    assert "Terra 5/6" in titel


def test_titel_erscheinen_nach_dem_abruf_in_der_tabelle(client):
    assert client.get("/api/rows").json()["cache_leer"] is True
    _abrufen(client)
    _warte_auf_ende(client)
    daten = client.get("/api/rows").json()
    assert daten["cache_leer"] is False
    zeile = next(z for z in daten["zeilen"] if z["bestand_ref"] == "G3")
    assert zeile["titel"] == "Terra 5/6"
    assert zeile["isbn"] == "978-3-12-105207-3"


def test_backup_wird_angelegt(client, workbook_path: Path):
    _abrufen(client)
    stand = _warte_auf_ende(client)
    name = stand["zusammenfassung"]["backup"]
    assert name and (workbook_path.parent / "backups" / name).is_file()


@pytest.fixture()
def client_leer(leeres_workbook: Path) -> TestClient:
    """Eine App auf der noch ungefüllten Mappe - der erste Abruf überhaupt.

    Die Fixture ``client`` arbeitet auf einer Mappe, in der der Snapshot schon
    steht (siehe ``conftest.workbook_path``); ein Abruf dort ändert
    definitionsgemäß nichts mehr. Für die Frage, was ein Abruf *ändert*, braucht
    es die leere Ausgangslage.
    """
    from app.settings import Einstellungen

    application = create_app(einstellungen=Einstellungen(
        iserv_domain="beispiel-schule.de",
        excel_pfad_kandidaten=(leeres_workbook,),
        blatt_raster=SHEET_NAME,
    ))
    with TestClient(application, base_url=TEST_BASIS_URL) as testclient:
        yield testclient


def test_zusammenfassung_nennt_die_wirklich_geaenderten_zellen(client_leer):
    """``geaenderte_refs`` ist die Grundlage der Hervorhebung in der Oberfläche.

    Der Browser verliert sein "vorher" beim Neuladen nach dem Abruf; die Bezüge
    kommen deshalb vom Server.
    """
    _abrufen(client_leer)
    stand = _warte_auf_ende(client_leer)
    assert stand["fehler"] is None, stand
    z = stand["zusammenfassung"]

    refs = z["geaenderte_refs"]
    assert "G3" in refs                       # war leer, steht jetzt auf 60
    assert len(refs) == z["geaendert"]
    assert z["geschrieben"] >= z["geaendert"]
    # Kein Bezug doppelt - sonst hübe die Oberfläche eine Zelle zweimal hervor
    # und entfernte die Marke beim ersten Ablauf wieder.
    assert len(set(refs)) == len(refs)


def test_ein_zweiter_abruf_meldet_nur_noch_den_stand_als_geaendert(client_leer):
    """Zweimal dasselbe abrufen heißt: außer der Uhrzeit hat sich nichts bewegt.

    ``UpdateResult.changes`` enthält jede geschriebene Zelle, auch die, in der
    schon dieselbe Zahl stand - würde die Oberfläche daran hängen, leuchtete
    nach jedem Abruf die ganze Tabelle auf und sagte damit nichts mehr.
    """
    _abrufen(client_leer)
    _warte_auf_ende(client_leer)
    _abrufen(client_leer)
    zweiter = _warte_auf_ende(client_leer)

    assert zweiter["fehler"] is None, zweiter
    z = zweiter["zusammenfassung"]
    assert z["geschrieben"] > 0
    # Keine einzige Zahl hat sich bewegt. Übrig bleiben kann allein die
    # "Stand"-Zelle B10 (der Abfragezeitpunkt) - und auch die nur, wenn die
    # beiden Läufe in verschiedene Sekunden fallen: der Zeitstempel wird auf
    # Sekunden gekürzt. Deshalb Teilmenge und nicht Gleichheit; eine
    # Gleichheitsprüfung wäre von der Laufzeit der Testsuite abhängig.
    assert set(z["geaenderte_refs"]) <= {"B10"}, z["geaenderte_refs"]


# ── Fehlerabbildung ───────────────────────────────────────────────────────────

def test_falsches_passwort_ist_401(client):
    from ausleihe.exceptions import AuthError

    antwort = _abrufen(client, _FehlerClient(AuthError("401")))
    assert antwort.status_code == 401
    assert "Zugangsdaten" in antwort.json()["fehler"]
    assert _manager(client).laeuft() is False


def test_fehlende_rolle_ist_403(client):
    from ausleihe.exceptions import ForbiddenError

    antwort = _abrufen(client, _FehlerClient(ForbiddenError("403")))
    assert antwort.status_code == 403
    assert "Ausleihe-Verwalter" in antwort.json()["fehler"]


def test_netzfehler_ist_504(client):
    from ausleihe.exceptions import TransportError

    antwort = _abrufen(client, _FehlerClient(TransportError("timeout")))
    assert antwort.status_code == 504
    assert "IServ" in antwort.json()["fehler"]


def test_unerwarteter_fehler_ist_500(client):
    antwort = _abrufen(client, _FehlerClient(RuntimeError("kaputt")))
    assert antwort.status_code == 500


@pytest.mark.parametrize("nutzlast", [
    {"benutzer": "", "passwort": PASSWORT},
    {"benutzer": BENUTZER, "passwort": ""},
    {"benutzer": BENUTZER},
    {},
])
def test_fehlende_zugangsdaten_sind_400(client, nutzlast):
    client.app.state.client_factory = FakeClient
    antwort = client.post("/api/refresh", json=nutzlast)
    assert antwort.status_code == 400


def test_mehrdeutige_zuordnung_speichert_nichts(client, workbook_path: Path):
    """Diagnosen -> 422 im Status, Mappe unverändert."""
    wb = load_workbook(str(workbook_path))
    wb["bestellt"].append([9, 2, None, "Terra 5/6", "Klett", "978-3-12-105207-3", 25.0])
    wb.save(str(workbook_path))
    vorher = workbook_path.stat().st_mtime

    _abrufen(client)
    stand = _warte_auf_ende(client)
    assert stand["fehlercode"] == 422
    assert stand["diagnosen"]
    assert "nichts gespeichert" in stand["fehler"]
    assert workbook_path.stat().st_mtime == vorher


def test_gesperrte_datei_ist_423(client, monkeypatch):
    from app import excel as excel_modul

    def verweigern(*args, **kwargs):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(excel_modul, "atomic_save_workbook", verweigern)
    _abrufen(client)
    stand = _warte_auf_ende(client)
    assert stand["fehlercode"] == 423
    assert "in Excel geöffnet" in stand["fehler"]


# ── Genau ein Lauf ────────────────────────────────────────────────────────────

def test_zweiter_abruf_waehrend_eines_laufenden_ist_409(client):
    _BlockierenderClient.freigabe.clear()
    try:
        erst = _abrufen(client, _BlockierenderClient)
        assert erst.status_code == 202
        # Warten, bis der Thread wirklich in der Serienabfrage hängt.
        frist = time.monotonic() + 5
        while time.monotonic() < frist and not _manager(client).laeuft():
            time.sleep(0.01)
        assert _manager(client).laeuft()

        zweit = _abrufen(client, FakeClient)
        assert zweit.status_code == 409
        assert "bereits ein Abruf" in zweit.json()["fehler"]
    finally:
        _BlockierenderClient.freigabe.set()
    _warte_auf_ende(client)


def test_nach_dem_lauf_geht_ein_neuer(client):
    assert _abrufen(client).status_code == 202
    _warte_auf_ende(client)
    assert _abrufen(client).status_code == 202
    assert _warte_auf_ende(client)["fehler"] is None


def test_status_ohne_lauf(client):
    stand = client.get("/api/refresh/status").json()
    assert stand["laeuft"] is False and stand["fertig"] is False
    assert stand["zusammenfassung"] is None
    assert stand["text"] == "Noch kein Abruf in dieser Sitzung."


def test_status_ohne_lauf_hat_dieselbe_schluesselmenge_wie_ein_echter_lauf(client):
    """Wächst ``Lauf`` um ein Feld, muss ``Lauf.ohne_lauf()`` automatisch mitziehen.

    Die erwartete Schlüsselmenge wird bewusst NICHT als Namensliste
    hingeschrieben (das wäre dieselbe Verdopplung nur eine Ebene tiefer),
    sondern aus den Dataclass-Feldern von ``Lauf`` selbst abgeleitet - und aus
    einem echten, per Abruf entstandenen Stand, damit auch tatsächlich das
    Web-Layer-Format geprüft wird und nicht nur das Dataclass-Innenleben.
    """
    erwartete_schluessel = {f.name for f in fields(Lauf)}
    # als_dict() ist die Übersetzung Dataclass -> JSON-Stand; beide Seiten
    # müssen exakt dieselben Schlüssel tragen, sonst redet der Test an der
    # eigentlichen Garantie vorbei.
    #
    # ⚠️ Wer dieses Feld später um ein rein internes ergänzt (Thread-Handle,
    # Zähler, irgendwann vielleicht doch ein Token), bekommt hier einen
    # Fehlschlag - und die bequeme Reparatur wäre, es in als_dict() zu
    # serialisieren und damit über /api/refresh/status nach außen zu geben.
    # Genau das darf nicht passieren: Lauf trägt laut eigenem Docstring
    # bewusst keine Zugangsdaten. Ein internes Feld gehört stattdessen hier
    # ausdrücklich ausgenommen (erwartete_schluessel - {"name"}), damit die
    # Entscheidung sichtbar getroffen und nicht stillschweigend weggetestet wird.
    assert set(Lauf.ohne_lauf().als_dict()) == erwartete_schluessel

    _abrufen(client)
    echter_lauf_stand = _warte_auf_ende(client)

    ohne_lauf_stand = RefreshManager().status()
    assert set(ohne_lauf_stand) == set(echter_lauf_stand) == erwartete_schluessel


# ── Zugangsdaten ──────────────────────────────────────────────────────────────

def test_passwort_taucht_in_keiner_antwort_auf(client, caplog):
    caplog.set_level(logging.DEBUG)
    antworten = [_abrufen(client)]
    _warte_auf_ende(client)
    antworten.append(client.get("/api/refresh/status"))
    antworten.append(client.get("/api/rows"))
    antworten.append(client.get("/"))
    for antwort in antworten:
        assert PASSWORT not in antwort.text
        assert BENUTZER not in antwort.text

    protokoll = "\n".join(eintrag.getMessage() for eintrag in caplog.records)
    assert PASSWORT not in protokoll


def test_passwort_steht_in_keinem_zustand(client, workbook_path: Path):
    _abrufen(client)
    _warte_auf_ende(client)

    assert PASSWORT not in json.dumps(_manager(client).status(), default=str)
    assert PASSWORT not in repr(vars(client.app.state))
    assert PASSWORT not in cache_modul.cache_pfad(workbook_path).read_text(encoding="utf-8")
    assert PASSWORT not in workbook_path.read_bytes().decode("latin-1")


def test_der_lauf_haelt_keine_zugangsdaten(client):
    """``Lauf`` ist das einzige Objekt, das den Abruf überlebt - es kennt sie nicht."""
    _abrufen(client)
    _warte_auf_ende(client)
    manager = _manager(client)
    with manager._zustand_lock:
        lauf = manager._aktueller
    assert PASSWORT not in repr(vars(lauf))


# ── Fortschritt ───────────────────────────────────────────────────────────────

def test_jahrgaenge_werden_einzeln_gemeldet():
    """Der längste Abschnitt bewegt den Balken - sonst stünde er eine Minute still."""
    from bestand.core import fetch_snapshot

    verlauf = []

    def merken(**felder):
        verlauf.append((felder["fortschritt"], felder["text"]))

    snapshot = fetch_snapshot(FakeClient(), "2026/2027")
    manager = RefreshManager()
    original = manager._setze
    manager._setze = merken
    try:
        geladen = manager._lade_jahrgaenge(snapshot)
    finally:
        manager._setze = original

    assert sorted(geladen) == [5, 6, 7, 12]
    assert [text for _, text in verlauf] == [
        f"Bücherliste Jahrgang {g}" for g in (5, 6, 7, 12)
    ]
    prozent = [wert for wert, _ in verlauf]
    assert prozent == sorted(prozent) and prozent[0] == 60 and prozent[-1] < 92


def test_fehlende_buecherliste_wird_zur_warnung(client, workbook_path: Path):
    """Ein Jahrgang ohne Bücherliste bricht nichts ab, steht aber im Bericht."""
    class _OhneJg7(FakeClient):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            echte = self.schoolyears

            class _Gefiltert:
                def get_current(self_inner):
                    return echte.get_current()

                def get_booklists(self_inner, sy_id):
                    return [bl for bl in echte.get_booklists(sy_id) if bl["grade"] != 7]

                def get_booklist(self_inner, sy_id, bl_id):
                    return echte.get_booklist(sy_id, bl_id)

            self.schoolyears = _Gefiltert()

    _abrufen(client, _OhneJg7)
    stand = _warte_auf_ende(client)
    assert stand["fehler"] is None, stand
    assert any("Jahrgang 7" in w for w in stand["warnungen"]), stand["warnungen"]
    # Jahrgang 7 steht im Raster, hat aber keine Liste -> die Zellen bleiben leer.
    zeilen = client.get("/api/rows").json()["zeilen"]
    jg7 = [z for z in zeilen if z["jahrgang"] == "7"]
    assert jg7 and all(z["titel"] is None for z in jg7)
