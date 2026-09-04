"""rows.py: Jahrgangsspanne, Angemeldet-Summe, Bedarf in Python gerechnet."""
from __future__ import annotations

from datetime import datetime

import pytest
from bestand.core import parse_grid
from bestand.core.testing import SHEET_NAME
from openpyxl import load_workbook

from app.cache import Cache, Eintrag
from app.rows import baue_zeilen


@pytest.fixture()
def zeilen(workbook_path):
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    return ws, baue_zeilen(ws, parse_grid(ws))


def _finde(zeilen, ref):
    return next(z for z in zeilen if z.bestand_ref == ref)


def test_mehrjahresband_wird_als_spanne_angezeigt(zeilen):
    _, liste = zeilen
    band = _finde(liste, "G3")
    assert band.jahrgang == "5-6"
    assert band.jahrgaenge == (5, 6)
    assert band.angemeldet_refs == ("F3", "F4")


def test_angemeldet_wird_ueber_die_jahrgaenge_summiert(zeilen):
    """48 (Jg 5) + 44 (Jg 6) - genau das, was die Excel-Formel =F3+F4-G3-H3 tut."""
    _, liste = zeilen
    assert _finde(liste, "G3").angemeldet == 92


def test_einzeljahrgang_bleibt_einzeln(zeilen):
    _, liste = zeilen
    zeile = _finde(liste, "C3")
    assert zeile.jahrgang == "5"
    assert zeile.angemeldet == 50


def test_bedarf_wird_in_python_gerechnet(zeilen):
    """92 angemeldet - 60 Bestand - 30 bestellt = 2. Ohne Sicherheitsbestand."""
    ws, liste = zeilen
    band = _finde(liste, "G3")
    assert (band.bestand, band.bestellt, band.zu_bestellen) == (60, 30, 2)
    assert band.bedarf is True
    # Die Formelspalte bleibt Formeltext und wird nie als Wert gelesen.
    assert ws["I3"].value == "=F3+F4-G3-H3"


def test_leere_bestellt_zelle_zaehlt_als_null(zeilen):
    _, liste = zeilen
    zeile = _finde(liste, "C5")   # Deutsch Jg 7: 52 angemeldet, 90 Bestand, nichts bestellt
    assert zeile.bestellt is None
    assert zeile.zu_bestellen == -38
    assert zeile.bedarf is False


def test_formelzelle_wird_nicht_als_zahl_gelesen(leeres_workbook):
    """Ohne Abruf sind alle Zahlen leer - die Formeln dürfen nicht durchschlagen."""
    wb = load_workbook(str(leeres_workbook))
    ws = wb[SHEET_NAME]
    liste = baue_zeilen(ws, parse_grid(ws))
    assert all(z.angemeldet is None for z in liste)
    assert all(z.zu_bestellen is None for z in liste)


def test_sperrflaechen_erscheinen_nicht(zeilen):
    _, liste = zeilen
    assert not any(z.fach == "Latein" and 5 in z.jahrgaenge for z in liste)
    # Jg 7 ist keine Sperrfläche, dort fehlt nur das Buch.
    assert any(z.fach == "Latein" and z.jahrgaenge == (7,) for z in liste)


def test_titel_und_isbn_kommen_aus_dem_cache(workbook_path):
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    band = next(e for e in grid.entries if e.slots["bestand"].ref == "G3")
    cache = Cache(stand=datetime(2026, 9, 4, 12, 0),
                  eintraege={band.key: Eintrag(isbn="978-3-12-105207-3", titel="Terra 5/6",
                                               preis=25.0)})
    zeile = _finde(baue_zeilen(ws, grid, cache), "G3")
    assert (zeile.titel, zeile.isbn) == ("Terra 5/6", "978-3-12-105207-3")


def test_ohne_cache_bleiben_titel_leer(zeilen):
    _, liste = zeilen
    assert all(z.titel is None and z.isbn is None for z in liste)
